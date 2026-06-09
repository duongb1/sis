import argparse
import csv
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from sklearn.model_selection import StratifiedGroupKFold

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from sislib.data.labels import (
    EXCEL_MULTICLASS_LABELS,
    binary_i63_from_multiclass,
    major_icd_code,
    mri_3class_from_maicd,
)


DEFAULT_KAGGLE_MRI_ROOT = "/kaggle/input/datasets/duongbui/siscth/mri"


def parse_args():
    parser = argparse.ArgumentParser(description="Build a 3-class MRI manifest and 5-fold patient-group splits.")
    parser.add_argument("--image-root", default=f"{DEFAULT_KAGGLE_MRI_ROOT}/images")
    parser.add_argument("--excel", default=f"{DEFAULT_KAGGLE_MRI_ROOT}/mri_3class.xlsx")
    parser.add_argument("--co-excel", default=None)
    parser.add_argument("--khong-excel", default=None)
    parser.add_argument("--out-manifest", default="/kaggle/working/mri_3class_manifest.csv")
    parser.add_argument("--out-folds", default="/kaggle/working/mri_3class_folds.csv")
    parser.add_argument("--folds", type=int, default=5)
    parser.add_argument("--val-ratio", type=float, default=0.1)
    parser.add_argument("--seed", type=int, default=42)
    return parser.parse_args()


def read_excel_rows(path, source_label=""):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Final"]
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_to_index = {header: index for index, header in enumerate(headers) if header is not None}
    required = ["STT", "MABN", "NGAYYLENH", "MAICD"]
    missing = [name for name in required if name not in header_to_index]
    if missing:
        raise ValueError(f"{path} is missing columns: {', '.join(missing)}")

    rows = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        stt = values[header_to_index["STT"]]
        maicd = values[header_to_index["MAICD"]]
        label = mri_3class_from_maicd(maicd)
        if "LABEL" in header_to_index and values[header_to_index["LABEL"]]:
            label = str(values[header_to_index["LABEL"]]).strip()
        row_source_label = source_label
        if "SOURCE_BINARY_LABEL" in header_to_index and values[header_to_index["SOURCE_BINARY_LABEL"]]:
            row_source_label = str(values[header_to_index["SOURCE_BINARY_LABEL"]]).strip()
        rows.append(
            {
                "case_id": str(stt),
                "source_binary_label": row_source_label,
                "source_file": Path(path).name,
                "excel_row": excel_row,
                "STT": stt,
                "MABN": values[header_to_index["MABN"]],
                "NGAYYLENH": values[header_to_index["NGAYYLENH"]],
                "MAICD": maicd,
                "MAICD_MAJOR": major_icd_code(maicd),
                "label_3class": label,
                "label_3class_id": EXCEL_MULTICLASS_LABELS.index(label),
                "binary_i63": binary_i63_from_multiclass(label),
            }
        )
    return rows


def count_mri_files(case_dir):
    adc = len(list((case_dir / "ADC").glob("*.JPG"))) if (case_dir / "ADC").is_dir() else 0
    dwi = len(list((case_dir / "DWI").glob("*.JPG"))) if (case_dir / "DWI").is_dir() else 0
    other = len([path for path in case_dir.rglob("*.JPG") if path.parent.name not in {"ADC", "DWI"}])
    return adc, dwi, other


def attach_image_paths(rows, image_root):
    missing = []
    for row in rows:
        label_dir = "co" if row["source_binary_label"] == "co" else "khong"
        case_dir = image_root / row["case_id"]
        if not case_dir.is_dir():
            case_dir = image_root / label_dir / row["case_id"]
        if not case_dir.is_dir():
            missing.append(str(case_dir))
            continue
        adc, dwi, other = count_mri_files(case_dir)
        row["image_dir"] = str(case_dir)
        row["adc_images"] = adc
        row["dwi_images"] = dwi
        row["other_jpg_images"] = other
        row["total_jpg_images"] = adc + dwi + other
    if missing:
        raise FileNotFoundError("Missing image folders:\n" + "\n".join(missing[:20]))
    return rows


def assign_group_folds(rows, n_folds, val_ratio, seed):
    y = [row["label_3class"] for row in rows]
    groups = [str(row["MABN"] or row["case_id"]) for row in rows]
    splitter = StratifiedGroupKFold(n_splits=n_folds, shuffle=True, random_state=seed)

    fold_records = []
    for fold_index, (train_val_idx, test_idx) in enumerate(splitter.split(rows, y, groups)):
        train_val_idx = list(train_val_idx)
        test_idx = list(test_idx)
        val_size = val_ratio / (1.0 - 1.0 / n_folds)
        val_folds = max(2, round(1.0 / val_size))
        val_splitter = StratifiedGroupKFold(n_splits=val_folds, shuffle=True, random_state=seed + fold_index)
        train_val_y = [y[index] for index in train_val_idx]
        train_val_groups = [groups[index] for index in train_val_idx]
        inner_splits = list(val_splitter.split(train_val_idx, train_val_y, train_val_groups))
        inner_train_positions, val_positions = inner_splits[fold_index % val_folds]
        train_idx = [train_val_idx[position] for position in inner_train_positions]
        val_idx = [train_val_idx[position] for position in val_positions]
        train_groups = {groups[index] for index in train_idx}
        val_groups = {groups[index] for index in val_idx}
        test_groups = {groups[index] for index in test_idx}
        if train_groups & val_groups or train_groups & test_groups or val_groups & test_groups:
            raise RuntimeError(f"Group leakage detected in fold {fold_index}.")

        split_by_index = {index: "train" for index in train_idx}
        split_by_index.update({index: "val" for index in val_idx})
        split_by_index.update({index: "test" for index in test_idx})
        for index, row in enumerate(rows):
            output = dict(row)
            output["fold"] = fold_index
            output["split"] = split_by_index[index]
            fold_records.append(output)
    return fold_records


def write_csv(path, rows):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def print_distribution(title, rows):
    counter = Counter(row["label_3class"] for row in rows)
    total = sum(counter.values())
    print(title)
    for label in EXCEL_MULTICLASS_LABELS:
        count = counter[label]
        pct = (count / total * 100.0) if total else 0.0
        print(f"  {label}: {count} ({pct:.2f}%)")


def print_fold_distribution(fold_rows):
    by_fold_split = defaultdict(list)
    for row in fold_rows:
        by_fold_split[(row["fold"], row["split"])].append(row)
    for fold in sorted({row["fold"] for row in fold_rows}):
        print(f"fold {fold}")
        for split in ["train", "val", "test"]:
            print_distribution(f"  {split}", by_fold_split[(fold, split)])


def main():
    args = parse_args()
    image_root = Path(args.image_root)
    rows = []
    if args.co_excel or args.khong_excel:
        if not args.co_excel or not args.khong_excel:
            raise ValueError("--co-excel and --khong-excel must be provided together.")
        rows.extend(read_excel_rows(args.co_excel, "co"))
        rows.extend(read_excel_rows(args.khong_excel, "khong"))
    else:
        rows.extend(read_excel_rows(args.excel))
    rows = attach_image_paths(rows, image_root)
    fold_rows = assign_group_folds(rows, args.folds, args.val_ratio, args.seed)

    write_csv(args.out_manifest, rows)
    write_csv(args.out_folds, fold_rows)

    print(f"wrote {args.out_manifest}")
    print(f"wrote {args.out_folds}")
    print_distribution("overall", rows)
    print_fold_distribution(fold_rows)


if __name__ == "__main__":
    main()
